import os, glob
from shutil import copyfile
import numpy as np
import openpyxl
import pandas as pd
from colour import Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from args import Args

args = Args().parse()


class ExperimentResults:
    def __init__(self):
        self.args = args
        self.device = self.args.cuda_device
        self.results_path = os.path.join(self.args.resultsRoot, 'results.xlsx')
        self.sorted_results_path = os.path.join(self.args.resultsRoot, 'sorted_result.xlsx')
        self.query_results_path = os.path.join(self.args.resultsRoot, 'queried_result.xlsx')
        self.top_results_path = os.path.join(self.args.resultsRoot, 'top_result.xlsx')
        self.COLOUR = ['DC143C', '800080', '0000FF', '00FFFF', '3CB371', 'FFD700', 'FF8C00', '8B0000', '000000',
                       'FF00FF', 'D2691E', '808080']

    # Save experiment result into a excel file
    def save_res_to_excel(self, results_data):
        if not os.path.exists(self.results_path):
            results_data.to_excel(self.results_path, index=False)
            self.adjust_excel_cell(self.results_path)
        else:
            results = pd.read_excel(self.results_path, engine='openpyxl')
            new_results = results.append(results_data, ignore_index=False)
            new_results.to_excel(self.results_path, index=False)
            self.adjust_excel_cell(self.results_path)

    # Adjust the width of cell and center each cell for excel
    @staticmethod
    def adjust_excel_cell(file_path, *color_column, fill_color=False):
        wb = openpyxl.load_workbook(filename=file_path)
        worksheet = wb.active
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if not pd.isnull(cell.value):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except IOError as e:
                        print(e)
                        pass
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                if fill_color and len(color_column) != 0:
                    red = Color('#ff9900')
                    colors = list(red.range_to(Color("#ff99ff"), len(color_column) + 4))
                    for i in color_column:
                        whether_fill_color = True if i == col[0].value else False
                        if whether_fill_color:
                            color_number = colors[color_column.index(i)].hex_l.replace('#', '')
                            now_colour = PatternFill("solid", fgColor=color_number)
                            cell.fill = now_colour
            if column == 'O':
                adjusted_width = (max_length + 2) / 2
            else:
                adjusted_width = max_length + 2.8
            worksheet.column_dimensions[column].width = adjusted_width
        wb.save(file_path)

    # index: eg. {'Dataset': True, 'Teacher_acc':True}
    def sort_excel(self, in_file_path, out_filepath, *index_list, ascending=False):
        results = pd.read_excel(in_file_path, engine='openpyxl')
        indexes = []
        ascending_orders = [ascending for i in index_list]
        marked_indexes = []
        # print(index_list)
        for i in index_list:
            indexes.append(i)
            if i.find('acc') != -1 or i.find('Acc') != -1:
                results[i] = results[i].apply(lambda x: str(x).replace('%', '')).astype('float64')
                marked_indexes.append(i)
            if i.find('Duration') != -1:
                results[i] = results[i].apply(lambda x: str(x).replace('s', '')).astype('float64')
                marked_indexes.append(i)
            if i.find('H') != -1:
                results[i] = results[i].apply(lambda x: str(x)).astype('float64')
                marked_indexes.append(i)
        results.sort_values(by=indexes, ascending=ascending_orders, inplace=True)
        for index in marked_indexes:
            results[index] = results[index].astype('object').apply(
                lambda x: str(x).replace('nan', '') if str(x) == 'nan' else str(x) + '%')
            if index.find('Duration') != -1:
                results[index] = results[index].apply(lambda x: str(x).replace('%', '') + 's')
            if index.find('H') != -1:
                results[index] = results[index].apply(lambda x: str(x).replace('%', ''))
        results.to_excel(out_filepath, index=False)
        self.adjust_excel_cell(out_filepath, *indexes, fill_color=True)

    def query_by(self, in_result_path):
        query_dict = self.args.query_keywords
        column_name, row_name = query_dict.get('keyword'), query_dict.get('filter_by')
        results = pd.read_excel(in_result_path, engine='openpyxl')
        results.set_index(column_name, inplace=True)
        query_result = results.loc[[row_name]]
        query_result.to_excel(self.query_results_path, index=True)
        if column_name not in set(self.args.sorted_by):
            self.args.sorted_by.insert(0, column_name)
        self.sort_excel(self.query_results_path, self.query_results_path, *self.args.sorted_by,
                        ascending=self.args.Ascending)

    def query_top_results(self, top_num=5):
        top_results = pd.DataFrame(columns=self.args.column_names)
        for name in self.args.all_dataset_names:
            top_res = self.check_best_model(name, top_num=top_num, print_results=False)
            top_results = top_results.append(top_res, ignore_index=False)
        top_results.to_excel(self.top_results_path, index=False)
        self.adjust_excel_cell(self.top_results_path)

    def check_best_model(self, dataset, top_num=5, print_results=True):
        best_result = 0
        self.args.query_keywords = {"keyword": "Dataset", "filter_by": dataset}
        query_details = [self.args.framework, self.args.task_categories, self.args.AZSL_test] if print_results else [
            self.args.q_fw, self.args.q_task_cg, self.args.q_AZSL_test]
        if not os.path.exists(self.results_path):
            best_result = None
        else:
            try:
                self.query_by(self.results_path)
                results = pd.read_excel(self.query_results_path, engine='openpyxl')
                results_head = None
                if query_details[1] == 'GZSL_all':
                    results_head = results.loc[
                        (results.Framework == query_details[0]) & (
                                results.Task_categories == query_details[1])]
                    best_result = results_head.head(top_num).H.values[0]
                elif query_details[1] == 'AZSL':
                    if query_details[2] == 'zsl':
                        results_head = results.loc[
                            (results.Framework == query_details[0]) & (
                                    results.Task_categories == query_details[1]) & (
                                    results.AZSL_test == query_details[2])]
                        best_result = results_head.head(top_num).Acc_unseen.values[0]
                    elif query_details[2] == 'gzsl':
                        results_head = results.loc[
                            (results.Framework == query_details[0]) & (
                                    results.Task_categories == query_details[1]) & (
                                    results.AZSL_test == query_details[2])]
                        best_result = results_head.head(top_num).H.values[0]
                self.set_pd_options()
                if print_results:
                    print('\n', '*' * 40,
                          ' | Previous Top5: {} / {} / {} | '.format(self.args.framework, self.args.task_categories,
                                                                     self.args.dataset),
                          '*' * 160, '\n', results_head.head(top_num)[self.args.prompt_info],
                          '\n', '\n', '*' * 40, ' | Model Details | ', '*' * 200, '\n')
                best_result = float(str(best_result).replace('%', '')) if str(best_result).find('%') else best_result
            except (ZeroDivisionError, Exception) as e:
                best_result = None
        if print_results:
            return best_result
        else:
            best_result = results_head.head(top_num)
            return best_result

    @staticmethod
    def move_best_model(input_dir, out_dir, file_suffix, current_result):
        original_file = glob.glob(os.path.join(input_dir, '*{}'.format(file_suffix)))
        for item in original_file:
            original_file_name = item.split('/')[-1]
            copyfile(item, out_dir + '/' + current_result[0] + '@' + str(current_result[1]) + '@' + original_file_name)
        print('Saved the best model in {}'.format(out_dir))

    # Generate random color
    @staticmethod
    def random_color(color_num):
        color = []
        int_num = [str(x) for x in np.arange(10)]
        alphabet = [chr(x) for x in (np.arange(6) + ord('A'))]  # Out[139]: ['A', 'B', 'C', 'D', 'E', 'F']
        color_arr = np.hstack((int_num, alphabet))
        for j in range(color_num):
            color_single = ''  # if need '#', replace '' by '#'
            for i in range(6):
                index = np.random.randint(len(color_arr))
                color_single += color_arr[index]
            color.append(color_single)
        return color  # Out[150]: ['#81D4D4', '#70344F', '#DF91B1', '#7EE250', '#C47BC3', '#9F88D5']

    # Convert hex_color to rgb/rgba color
    @staticmethod
    def hex_to_rgb(hex_color):
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        rgb = [r, g, b]  # rgb color to rgba color
        return rgb

    @staticmethod
    def set_pd_options():
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 1000)
        pd.set_option('max_colwidth', 1000)
        pd.set_option('display.unicode.ambiguous_as_wide', True)



if __name__ == '__main__':
    exp_res = ExperimentResults()
    if args.check_result:
        if args.query_type == 'top_all_dataset':
            exp_res.query_top_results(top_num=args.check_top_num)
            print('\n', '*' * 100, '\n',
                  'Query top {} performance in {} / {} setting , please check "{}" !'.format(args.check_top_num,
                                                                                             args.q_fw, args.q_task_cg,
                                                                                             exp_res.top_results_path),
                  '\n', '*' * 100)
        elif args.query_type == 'sort_all_dataset':
            exp_res.sort_excel(exp_res.results_path, exp_res.sorted_results_path, *args.sorted_by,
                               ascending=args.Ascending)
            print('\n', '*' * 100, '\n',
                  'Sorted all previous results, please check "{}"'.format(exp_res.sorted_results_path),
                  '\n', '*' * 100)
        elif args.query_type == 'query_one_dataset':
            exp_res.query_by(exp_res.results_path)
            dataset_name = args.query_keywords.get('filter_by')
            print('\n', '*' * 100, '\n',
                  'Complete query the result in {} dataset, please check "{}" !'.format(dataset_name,
                                                                                        exp_res.sorted_results_path),
                  '\n', '*' * 100)
